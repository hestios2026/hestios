"""
Tagesbericht API — mobile field reporting endpoints.
POST /tagesbericht/          — create entry (sync from mobile)
POST /tagesbericht/photos/   — upload photo for an entry
GET  /tagesbericht/          — list entries (filter by site_id, date)
GET  /tagesbericht/{id}/     — get single entry with photos
"""
from __future__ import annotations

import io
import uuid
import logging
from datetime import date, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.auth import get_current_user
from app.core.database import get_db
from app.models.tagesbericht import TagesberichtEntry, TagesberichtPhoto
from app.models.user import User
from app.models.site import Site

log = logging.getLogger(__name__)

try:
    from app.core.storage import upload_file, get_presigned_url
    STORAGE_AVAILABLE = True
except ImportError:
    STORAGE_AVAILABLE = False

router = APIRouter(prefix="/tagesbericht", tags=["tagesbericht"])


def _fresh_url(s3_key: str) -> str:
    """Generate a fresh 7-day presigned URL for a given S3 key."""
    if STORAGE_AVAILABLE and s3_key and not s3_key.startswith("http"):
        try:
            return get_presigned_url(s3_key, expires_seconds=86400 * 7)
        except Exception:
            pass
    return s3_key


# ─── Schemas ──────────────────────────────────────────────────────────────────

class EntryIn(BaseModel):
    id: str                   # local UUID from mobile
    site_id: int
    nvt_number: str = ""
    work_type: str
    created_by: int
    created_by_name: str = ""
    created_at: str           # ISO string
    data: dict


class EntryOut(BaseModel):
    id: int
    local_uuid: str
    site_id: int
    work_type: str
    nvt_number: str
    created_by: int
    created_at: str
    synced_at: str
    data: dict
    photos: list[dict] = []

    class Config:
        from_attributes = True


# ─── Routes ───────────────────────────────────────────────────────────────────

@router.post("/", status_code=201)
def create_entry(
    body: EntryIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Idempotent — skip if already synced
    existing = db.query(TagesberichtEntry).filter_by(local_uuid=body.id).first()
    if existing:
        return {"id": existing.id, "status": "already_synced"}

    # Strip photos from data before storing (photos uploaded separately)
    data_clean = {k: v for k, v in body.data.items() if k != "photos"}

    try:
        created_at = datetime.fromisoformat(body.created_at.replace("Z", "+00:00"))
    except Exception:
        created_at = datetime.utcnow()

    entry = TagesberichtEntry(
        local_uuid=body.id,
        site_id=body.site_id,
        work_type=body.work_type,
        nvt_number=body.nvt_number,
        created_by=body.created_by,
        created_at=created_at,
        data=data_clean,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "status": "created"}


@router.post("/photos/")
async def upload_photo(
    entry_id: int = Form(...),
    category: str = Form(default="Altele"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    entry = db.query(TagesberichtEntry).filter_by(id=entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")

    content = await file.read()
    filename = f"tagesbericht/{entry.site_id}/{entry.work_type}/{entry_id}_{uuid.uuid4().hex[:8]}.jpg"
    url = filename  # fallback if storage unavailable

    if STORAGE_AVAILABLE:
        try:
            upload_file(filename, content, content_type="image/jpeg")
            url = get_presigned_url(filename, expires_seconds=86400 * 7)  # 7 days
        except Exception:
            pass

    photo = TagesberichtPhoto(
        entry_id=entry_id,
        category=category,
        filename=file.filename or filename,
        s3_key=filename,
        url=url,
        file_size=len(content),
    )
    db.add(photo)
    db.commit()
    db.refresh(photo)
    return {"id": photo.id, "url": url}


@router.get("/")
def list_entries(
    site_id: Optional[int] = Query(None),
    work_type: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(TagesberichtEntry)
    if site_id:
        q = q.filter(TagesberichtEntry.site_id == site_id)
    if work_type:
        q = q.filter(TagesberichtEntry.work_type == work_type)
    if date_from:
        q = q.filter(TagesberichtEntry.created_at >= date_from)
    if date_to:
        q = q.filter(TagesberichtEntry.created_at <= date_to)
    entries = q.order_by(TagesberichtEntry.created_at.desc()).limit(limit).all()

    result = []
    for e in entries:
        photos = db.query(TagesberichtPhoto).filter_by(entry_id=e.id).all()
        result.append({
            "id": e.id,
            "local_uuid": e.local_uuid,
            "site_id": e.site_id,
            "work_type": e.work_type,
            "nvt_number": e.nvt_number,
            "created_by": e.created_by,
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "synced_at": e.synced_at.isoformat() if e.synced_at else None,
            "data": e.data,
            "photos": [{"id": p.id, "category": p.category, "url": _fresh_url(p.s3_key), "filename": p.filename} for p in photos],
        })
    return result


# ─── Export helpers ───────────────────────────────────────────────────────────

WORK_TYPE_LABELS = {
    'poze_inainte': 'Poze Înainte', 'teratest': 'Teratest',
    'semne_circulatie': 'Semne Circulație', 'liefer_scheine': 'Liefer Scheine',
    'montaj_nvt_pdp': 'Montaj NVT/PDP', 'hp_plus': 'HP+', 'ha': 'HA',
    'reparatie': 'Reparație', 'tras_teava': 'Tras Țeavă', 'groapa': 'Groapă',
    'traversare': 'Traversare', 'sapatura': 'Săpătură',
    'raport_zilnic': 'Raport Zilnic', 'comanda_materiale': 'Comandă Materiale',
}

DATA_FIELD_LABELS = {
    'locatie': 'Locație', 'locatie_start': 'Start', 'locatie_stop': 'Stop',
    'start': 'Start', 'stop': 'Stop',
    'tip_conectare': 'Tip Conectare', 'suprafata': 'Suprafață',
    'suprafata_mixt_detalii': 'Detalii Mixt', 'lungime': 'Lungime (m)',
    'terasament': 'Terasament', 'grosime_asfalt': 'Grosime Asfalt (cm)',
    'latime': 'Lățime (m)', 'adancime': 'Adâncime (m)', 'tip': 'Tip Săpătură',
    'nr_cabluri': 'Nr. Cabluri', 'teava_protectie': 'Țeavă Protecție',
    'lungime_totala': 'Lungime Totală (m)', 'nr_bransamente_ha': 'Nr. Branșamente HA',
    'nr_hp_plus': 'Nr. HP+', 'moment': 'Moment', 'descriere': 'Descriere',
    'materiale': 'Materiale', 'urgenta': 'Urgență', 'notes': 'Observații',
    'comentarii': 'Comentarii',
}

SKIP_FIELDS = {'photos', 'waypoints'}


def _parse_coord(v: str):
    """Parse 'Address | lat,lng' or 'lat,lng' → (lat, lng) or None"""
    if not v or not isinstance(v, str): return None
    s = v.split('|')[-1].strip()
    parts = [p.strip() for p in s.split(',')]
    if len(parts) == 2:
        try:
            lat, lng = float(parts[0]), float(parts[1])
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                return (lat, lng)
        except ValueError:
            pass
    return None


def _render_map(data: dict, width=640, height=400) -> Optional[bytes]:
    """Render a static map PNG for an entry's coordinates."""
    try:
        import PIL.Image
        if not hasattr(PIL.Image, 'ANTIALIAS'):
            PIL.Image.ANTIALIAS = PIL.Image.LANCZOS  # removed in Pillow 10
        from staticmap import StaticMap, CircleMarker, Line
        start  = _parse_coord(str(data.get('start') or data.get('locatie_start') or ''))
        stop   = _parse_coord(str(data.get('stop')  or data.get('locatie_stop')  or ''))
        single = _parse_coord(str(data.get('locatie') or ''))
        wps_raw = data.get('waypoints') or []
        waypoints = [_parse_coord(str(w)) for w in wps_raw if _parse_coord(str(w))]

        if not start and not stop and not single:
            return None

        m = StaticMap(width, height,
                      url_template='https://a.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}.png',
                      headers={'User-Agent': 'HestiOS-Export/1.0'})

        if single:
            m.add_marker(CircleMarker((single[1], single[0]), '#22c55e', 16))
        if start:
            m.add_marker(CircleMarker((start[1], start[0]), '#22c55e', 16))
        for wp in waypoints:
            m.add_marker(CircleMarker((wp[1], wp[0]), '#f97316', 12))
        if stop:
            m.add_marker(CircleMarker((stop[1], stop[0]), '#ef4444', 16))

        route = [p for p in [start, *waypoints, stop] if p]
        if len(route) >= 2:
            m.add_line(Line([(p[1], p[0]) for p in route], '#f97316', 3))

        zoom = 17
        img = m.render(zoom=zoom)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        return buf.getvalue()
    except Exception as e:
        log.warning("Map render failed: %s", e)
        return None


def _entry_to_row(e: TagesberichtEntry, site_name: str, photo_count: int) -> dict:
    d = e.data or {}
    row = {
        'ID': e.id,
        'Data': e.created_at.strftime('%d.%m.%Y %H:%M') if e.created_at else '',
        'Tip Lucrare': WORK_TYPE_LABELS.get(e.work_type, e.work_type),
        'Șantier': site_name,
        'NVT': e.nvt_number or '',
    }
    for k, v in d.items():
        if k in SKIP_FIELDS: continue
        if isinstance(v, list): v = ', '.join(str(x) for x in v)
        row[DATA_FIELD_LABELS.get(k, k)] = str(v) if v else ''
    row['Nr. Fotografii'] = photo_count
    return row


# ─── Excel export ──────────────────────────────────────────────────────────────

def _build_excel(entries_data: list) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlsxImage
    from openpyxl.utils import get_column_letter

    GREEN = 'FF22C55E'
    DARK  = 'FF0C0F16'
    LIGHT = 'FFF0FDF4'
    GRAY  = 'FFF8FAFC'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rapoarte Mobile'

    # Collect all keys
    all_keys: list[str] = []
    for item in entries_data:
        for k in item['row'].keys():
            if k not in all_keys:
                all_keys.append(k)

    # Header
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = PatternFill('solid', fgColor=DARK)
        cell.font = Font(bold=True, color='FFFFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, item in enumerate(entries_data, 2):
        fill = PatternFill('solid', fgColor=LIGHT if row_idx % 2 == 0 else GRAY)
        for col_idx, key in enumerate(all_keys, 1):
            val = item['row'].get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.font = Font(size=10)

    # Maps & photos — separate sheet
    map_ws = wb.create_sheet('Hărți & Poze')
    map_row = 1
    for item in entries_data:
        has_map = bool(item.get('map_png'))
        photos = item.get('photos', [])
        if not has_map and not photos:
            continue

        # Entry header
        map_ws.cell(row=map_row, column=1,
                    value=f"#{item['row']['ID']} — {item['row']['Tip Lucrare']} — {item['row']['Data']}")
        map_ws.cell(row=map_row, column=1).font = Font(bold=True, size=11)
        map_ws.cell(row=map_row, column=1).fill = PatternFill('solid', fgColor=DARK)
        map_ws.cell(row=map_row, column=1).font = Font(bold=True, size=11, color='FFFFFFFF')
        map_row += 1

        if has_map:
            img_buf = io.BytesIO(item['map_png'])
            img = XlsxImage(img_buf)
            img.width, img.height = 640, 400
            map_ws.add_image(img, f'A{map_row}')
            map_ws.row_dimensions[map_row].height = 300
            map_row += 22

        # Embed photos (max 20 per entry to stay reasonable)
        for photo in photos[:20]:
            if not photo.get('bytes'):
                continue
            try:
                map_ws.cell(row=map_row, column=1,
                            value=f"Foto: {photo['category']}")
                map_ws.cell(row=map_row, column=1).font = Font(italic=True, size=9, color='FF64748B')
                map_row += 1
                p_buf = io.BytesIO(photo['bytes'])
                p_img = XlsxImage(p_buf)
                p_img.width, p_img.height = 480, 360
                map_ws.add_image(p_img, f'A{map_row}')
                map_ws.row_dimensions[map_row].height = 270
                map_row += 19
            except Exception:
                pass

        map_row += 2  # spacing between entries

    # Auto-width main sheet columns
    for col_idx, key in enumerate(all_keys, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(key)), *(len(str(item['row'].get(key, ''))) for item in entries_data))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── PDF export ───────────────────────────────────────────────────────────────

def _build_pdf(entries_data: list, site_name_filter: str = '') -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, HRFlowable, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    W, H = A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    GREEN = colors.HexColor('#22C55E')
    DARK  = colors.HexColor('#0C0F16')
    CARD  = colors.HexColor('#F8FAFC')

    styles = getSampleStyleSheet()
    title_style  = ParagraphStyle('Title', fontSize=18, textColor=DARK, fontName='Helvetica-Bold', spaceAfter=4)
    sub_style    = ParagraphStyle('Sub',   fontSize=10, textColor=colors.grey, spaceAfter=12)
    h2_style     = ParagraphStyle('H2',    fontSize=13, textColor=GREEN, fontName='Helvetica-Bold', spaceBefore=6, spaceAfter=4)
    label_style  = ParagraphStyle('Lbl',   fontSize=8,  textColor=colors.grey, fontName='Helvetica-Bold')
    value_style  = ParagraphStyle('Val',   fontSize=10, textColor=DARK)

    story = []
    # Cover / title
    story.append(Paragraph('HestiOS — Rapoarte Mobile', title_style))
    subtitle = f"Export {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    if site_name_filter:
        subtitle += f" · {site_name_filter}"
    story.append(Paragraph(subtitle, sub_style))
    story.append(HRFlowable(width='100%', thickness=2, color=GREEN, spaceAfter=16))

    for item in entries_data:
        row = item['row']
        story.append(Paragraph(f"{row['Tip Lucrare']} — {row['Data']}", h2_style))

        # Info table
        info_fields = [('Șantier', row.get('Șantier', '')),
                       ('NVT',     row.get('NVT', '')),
                       ('ID',      str(row.get('ID', '')))]
        info_data = [[Paragraph(k, label_style), Paragraph(str(v), value_style)] for k, v in info_fields if v]

        # Data fields
        skip = {'ID', 'Data', 'Tip Lucrare', 'Șantier', 'NVT', 'Nr. Fotografii'}
        for k, v in row.items():
            if k in skip or not v: continue
            info_data.append([Paragraph(k, label_style), Paragraph(str(v), value_style)])

        if info_data:
            t = Table(info_data, colWidths=[45*mm, W - 30*mm - 45*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), CARD),
                ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CARD]),
                ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('INNERGRID', (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('LEFTPADDING', (0,0), (-1,-1), 8),
            ]))
            story.append(t)

        # Map image
        if item.get('map_png'):
            story.append(Spacer(1, 8*mm))
            story.append(Paragraph('Traseu pe Hartă', h2_style))
            map_img = RLImage(io.BytesIO(item['map_png']),
                              width=W - 30*mm,
                              height=(W - 30*mm) * 400 / 640)
            story.append(map_img)

        # Photos
        photos = item.get('photos', [])
        valid_photos = [p for p in photos if p.get('bytes')]
        if valid_photos:
            story.append(Spacer(1, 4*mm))
            story.append(Paragraph(f"Fotografii ({len(valid_photos)})", h2_style))
            for p in valid_photos[:20]:
                try:
                    p_img = RLImage(io.BytesIO(p['bytes']),
                                   width=W - 30*mm,
                                   height=(W - 30*mm) * 3 / 4)
                    story.append(p_img)
                    if p.get('category'):
                        story.append(Paragraph(p['category'], label_style))
                    story.append(Spacer(1, 3*mm))
                except Exception:
                    pass
        else:
            story.append(Paragraph(f"Fotografii: {row.get('Nr. Fotografii', 0)}", label_style))

        story.append(Spacer(1, 6*mm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#E2E8F0'), spaceAfter=8))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─── Export endpoint ───────────────────────────────────────────────────────────

def _fetch_photo_bytes(url: str) -> Optional[bytes]:
    """Download photo bytes from a URL. Returns None on failure."""
    try:
        import requests as req
        r = req.get(url, timeout=15)
        if r.status_code == 200:
            return r.content
    except Exception as e:
        log.warning("Photo fetch failed %s: %s", url, e)
    return None


@router.get("/export/")
def export_entries(
    format: str = Query('excel', regex='^(excel|pdf)$'),
    ids: Optional[str] = Query(None),          # comma-separated entry IDs
    site_id: Optional[int] = Query(None),
    work_type: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    sort_by: str = Query('created_at'),
    sort_order: str = Query('desc'),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(TagesberichtEntry)

    if ids:
        id_list = [int(x.strip()) for x in ids.split(',') if x.strip().isdigit()]
        q = q.filter(TagesberichtEntry.id.in_(id_list))
    else:
        if site_id:
            q = q.filter(TagesberichtEntry.site_id == site_id)
        if work_type:
            q = q.filter(TagesberichtEntry.work_type == work_type)
        if date_from:
            q = q.filter(TagesberichtEntry.created_at >= date_from)
        if date_to:
            q = q.filter(TagesberichtEntry.created_at <= date_to)

    col_map = {'created_at': TagesberichtEntry.created_at,
               'work_type':  TagesberichtEntry.work_type,
               'site_id':    TagesberichtEntry.site_id}
    sort_col = col_map.get(sort_by, TagesberichtEntry.created_at)
    if sort_order == 'asc':
        q = q.order_by(sort_col.asc())
    else:
        q = q.order_by(sort_col.desc())

    entries = q.limit(500).all()

    # Build site map
    site_map: dict[int, str] = {}
    for s in db.query(Site).all():
        site_map[s.id] = s.name

    # Resolve site name filter label for PDF
    site_name_filter = site_map.get(site_id, '') if site_id else ''

    entries_data = []
    for e in entries:
        photos = db.query(TagesberichtPhoto).filter_by(entry_id=e.id).all()
        # Build fresh URLs and fetch bytes for embedding
        photo_items = []
        for p in photos:
            fresh_url = _fresh_url(p.s3_key)
            photo_bytes = _fetch_photo_bytes(fresh_url)
            photo_items.append({'category': p.category or '', 'bytes': photo_bytes})
        row = _entry_to_row(e, site_map.get(e.site_id, str(e.site_id)), len(photos))
        map_png = _render_map(e.data or {})
        entries_data.append({'row': row, 'map_png': map_png, 'photos': photo_items})

    ts = datetime.now().strftime('%Y%m%d_%H%M')

    if format == 'excel':
        content = _build_excel(entries_data)
        filename = f'rapoarte_mobile_{ts}.xlsx'
        media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        content = _build_pdf(entries_data, site_name_filter)
        filename = f'rapoarte_mobile_{ts}.pdf'
        media_type = 'application/pdf'

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.get("/{entry_id}/")
def get_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    entry = db.query(TagesberichtEntry).filter_by(id=entry_id).first()
    if not entry:
        raise HTTPException(404, "Not found")
    photos = db.query(TagesberichtPhoto).filter_by(entry_id=entry_id).all()
    return {
        "id": entry.id,
        "local_uuid": entry.local_uuid,
        "site_id": entry.site_id,
        "work_type": entry.work_type,
        "nvt_number": entry.nvt_number,
        "created_by": entry.created_by,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "synced_at": entry.synced_at.isoformat() if entry.synced_at else None,
        "data": entry.data,
        "photos": [{"id": p.id, "category": p.category, "url": _fresh_url(p.s3_key), "filename": p.filename} for p in photos],
    }
