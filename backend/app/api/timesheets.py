from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime, timedelta
import io
import csv

from app.core.database import get_db
from app.models.timesheet import TimeEntry, LeaveRequest, PayrollRecord
from app.models.employee import Employee
from app.models.user import User
from app.api.auth import get_current_user

router = APIRouter(prefix="/timesheets", tags=["timesheets"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _iso_week(d: date) -> int:
    return d.isocalendar()[1]


def _business_days(d_from: date, d_to: date) -> float:
    """Count Mon–Fri days between two dates inclusive."""
    count = 0
    cur = d_from
    while cur <= d_to:
        if cur.weekday() < 5:
            count += 1
        cur += timedelta(days=1)
    return float(count)


def _entry_dict(e: TimeEntry):
    emp = e.employee
    return {
        "id": e.id,
        "employee_id": e.employee_id,
        "employee_name": f"{emp.vorname} {emp.nachname}" if emp else None,
        "site_id": e.site_id,
        "date": e.date,
        "hours_regular": e.hours_regular,
        "hours_overtime": e.hours_overtime,
        "hours_night": e.hours_night,
        "hours_holiday": e.hours_holiday,
        "entry_type": e.entry_type,
        "notes": e.notes,
        "source": e.source,
        "week_number": e.week_number,
    }


def _leave_dict(l: LeaveRequest):
    return {
        "id": l.id,
        "employee_id": l.employee_id,
        "employee_name": f"{l.employee.vorname} {l.employee.nachname}" if l.employee else None,
        "leave_type": l.leave_type,
        "date_from": l.date_from,
        "date_to": l.date_to,
        "days_count": l.days_count,
        "status": l.status,
        "notes": l.notes,
        "approved_by": l.approved_by,
        "approved_at": l.approved_at,
        "created_at": l.created_at,
    }


def _payroll_dict(p: PayrollRecord):
    emp = p.employee
    return {
        "id": p.id,
        "employee_id": p.employee_id,
        "employee_name": f"{emp.vorname} {emp.nachname}" if emp else None,
        "personalnummer": emp.personalnummer if emp else None,
        "year": p.year,
        "month": p.month,
        "hours_regular": p.hours_regular,
        "hours_overtime": p.hours_overtime,
        "hours_night": p.hours_night,
        "hours_holiday": p.hours_holiday,
        "hours_sick": p.hours_sick,
        "hours_vacation": p.hours_vacation,
        "days_worked": p.days_worked,
        "brutto_regular": p.brutto_regular,
        "brutto_overtime": p.brutto_overtime,
        "brutto_night": p.brutto_night,
        "brutto_bauzuschlag": p.brutto_bauzuschlag,
        "brutto_total": p.brutto_total,
        "ag_sv_anteil": p.ag_sv_anteil,
        "soka_bau": p.soka_bau,
        "total_employer_cost": p.total_employer_cost,
        "status": p.status,
        "notes": p.notes,
        "locked_at": p.locked_at,
    }


# ── Schemas ───────────────────────────────────────────────────────────────────

class TimeEntryCreate(BaseModel):
    employee_id: int
    site_id: Optional[int] = None
    date: date
    hours_regular: float = 8.0
    hours_overtime: float = 0.0
    hours_night: float = 0.0
    hours_holiday: float = 0.0
    entry_type: str = "work"
    notes: Optional[str] = None


class TimeEntryUpdate(BaseModel):
    hours_regular: Optional[float] = None
    hours_overtime: Optional[float] = None
    hours_night: Optional[float] = None
    hours_holiday: Optional[float] = None
    entry_type: Optional[str] = None
    notes: Optional[str] = None
    site_id: Optional[int] = None


class LeaveRequestCreate(BaseModel):
    employee_id: int
    leave_type: str
    date_from: date
    date_to: date
    notes: Optional[str] = None


class PayrollRecalcRequest(BaseModel):
    year: int
    month: int
    employee_ids: Optional[List[int]] = None  # None = all active


# ── Time Entries ──────────────────────────────────────────────────────────────

@router.get("/entries/")
def list_entries(
    employee_id: Optional[int] = Query(None),
    week: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(TimeEntry)
    if employee_id:
        q = q.filter(TimeEntry.employee_id == employee_id)
    if week and year:
        q = q.filter(TimeEntry.week_number == week, sqlfunc.extract('year', TimeEntry.date) == year)
    if date_from:
        q = q.filter(TimeEntry.date >= date_from)
    if date_to:
        q = q.filter(TimeEntry.date <= date_to)
    return [_entry_dict(e) for e in q.order_by(TimeEntry.date.desc(), TimeEntry.employee_id).all()]


@router.post("/entries/", status_code=201)
def create_entry(body: TimeEntryCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Upsert: if entry for same employee+date exists, update it
    existing = db.query(TimeEntry).filter(
        TimeEntry.employee_id == body.employee_id,
        TimeEntry.date == body.date,
    ).first()
    if existing:
        for k, v in body.model_dump(exclude_none=True).items():
            setattr(existing, k, v)
        existing.week_number = _iso_week(body.date)
        db.commit()
        db.refresh(existing)
        return _entry_dict(existing)

    entry = TimeEntry(
        **body.model_dump(),
        week_number=_iso_week(body.date),
        created_by=current_user.id,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return _entry_dict(entry)


@router.put("/entries/{entry_id}/")
def update_entry(entry_id: int, body: TimeEntryUpdate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    e = db.query(TimeEntry).filter(TimeEntry.id == entry_id).first()
    if not e:
        raise HTTPException(404, "Not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(e, k, v)
    db.commit()
    return _entry_dict(e)


@router.delete("/entries/{entry_id}/", status_code=204)
def delete_entry(entry_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    e = db.query(TimeEntry).filter(TimeEntry.id == entry_id).first()
    if not e:
        raise HTTPException(404, "Not found")
    db.delete(e)
    db.commit()


# ── Weekly summary ────────────────────────────────────────────────────────────

@router.get("/weekly-summary/")
def weekly_summary(
    week: int = Query(...),
    year: int = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Returns one row per active employee with their hours for the given ISO week."""
    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.nachname).all()
    result = []
    for emp in employees:
        entries = db.query(TimeEntry).filter(
            TimeEntry.employee_id == emp.id,
            TimeEntry.week_number == week,
            sqlfunc.extract('year', TimeEntry.date) == year,
        ).order_by(TimeEntry.date).all()

        daily = {e.date.isoweekday(): e for e in entries}  # 1=Mon..5=Fri
        total_regular  = sum(e.hours_regular for e in entries if e.entry_type == "work")
        total_overtime = sum(e.hours_overtime for e in entries)
        total_absent   = sum(1 for e in entries if e.entry_type in ("sick", "vacation", "absent"))

        result.append({
            "employee_id": emp.id,
            "employee_name": f"{emp.vorname} {emp.nachname}",
            "personalnummer": emp.personalnummer,
            "tariflohn": emp.tariflohn,
            "bauzuschlag": emp.bauzuschlag,
            "total_regular": total_regular,
            "total_overtime": total_overtime,
            "total_absent_days": total_absent,
            "total_hours": total_regular + total_overtime,
            "days": {
                str(dow): _entry_dict(e) if e else None
                for dow, e in [(1, daily.get(1)), (2, daily.get(2)), (3, daily.get(3)),
                               (4, daily.get(4)), (5, daily.get(5))]
            },
        })
    return result


# ── Import from Excel ─────────────────────────────────────────────────────────

@router.post("/import/")
async def import_timesheet(
    file: UploadFile = File(...),
    week: int = Query(...),
    year: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Import weekly timesheet from Excel.
    Expected columns: Personalnummer | Montag | Dienstag | Mittwoch | Donnerstag | Freitag | Überstunden | Typ
    OR: Personalnummer | Mon | Tue | Wed | Thu | Fri | OT | Type
    Each cell = hours (float) or absence code (U=Urlaub, K=Krank, F=Feiertag, A=Absent)
    """
    import openpyxl
    from datetime import date as date_cls

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    # Get Monday of the given ISO week
    jan4 = date_cls(year, 1, 4)
    monday = jan4 + timedelta(weeks=week - 1, days=-jan4.weekday())

    # Read header row to find column indices
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def col_idx(names):
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    pnr_col = col_idx(["personalnummer", "personal-nr", "nr", "person"])
    mon_col = col_idx(["montag", "mon", "mo"])
    tue_col = col_idx(["dienstag", "tue", "di"])
    wed_col = col_idx(["mittwoch", "wed", "mi"])
    thu_col = col_idx(["donnerstag", "thu", "do"])
    fri_col = col_idx(["freitag", "fri", "fr"])
    ot_col  = col_idx(["überstunden", "uberstunden", "overtime", "ot"])

    if pnr_col is None:
        raise HTTPException(400, "Coloana 'Personalnummer' nu a fost găsită în fișier")

    ABSENCE_MAP = {"u": "vacation", "k": "sick", "f": "holiday", "a": "absent",
                   "ur": "vacation", "kr": "sick", "fe": "holiday"}

    imported, skipped = 0, 0
    for row in range(2, ws.max_row + 1):
        pnr_val = ws.cell(row, pnr_col + 1).value
        if not pnr_val:
            continue
        pnr = str(pnr_val).strip()
        emp = db.query(Employee).filter(Employee.personalnummer == pnr).first()
        if not emp:
            skipped += 1
            continue

        day_cols = [mon_col, tue_col, wed_col, thu_col, fri_col]
        day_offsets = [0, 1, 2, 3, 4]

        overtime_total = 0.0
        if ot_col is not None:
            ot_val = ws.cell(row, ot_col + 1).value
            try:
                overtime_total = float(ot_val or 0)
            except (ValueError, TypeError):
                overtime_total = 0.0

        for i, dc in enumerate(day_cols):
            if dc is None:
                continue
            cell_val = ws.cell(row, dc + 1).value
            if cell_val is None:
                continue

            entry_date = monday + timedelta(days=day_offsets[i])
            entry_type = "work"
            hours_regular = 0.0
            hours_overtime = 0.0

            cell_str = str(cell_val).strip().lower()
            if cell_str in ABSENCE_MAP:
                entry_type = ABSENCE_MAP[cell_str]
                hours_regular = 0.0
            else:
                try:
                    hours_regular = float(cell_val)
                    entry_type = "work"
                except (ValueError, TypeError):
                    continue

            # Distribute overtime across worked days (simple: add to Friday)
            if i == 4 and overtime_total > 0:
                hours_overtime = overtime_total

            # Upsert
            existing = db.query(TimeEntry).filter(
                TimeEntry.employee_id == emp.id,
                TimeEntry.date == entry_date,
            ).first()
            if existing:
                existing.hours_regular = hours_regular
                existing.hours_overtime = hours_overtime
                existing.entry_type = entry_type
                existing.source = "import"
                existing.week_number = week
            else:
                db.add(TimeEntry(
                    employee_id=emp.id,
                    date=entry_date,
                    hours_regular=hours_regular,
                    hours_overtime=hours_overtime,
                    entry_type=entry_type,
                    source="import",
                    week_number=week,
                    created_by=current_user.id,
                ))
            imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}


# ── Leave requests ────────────────────────────────────────────────────────────

@router.get("/leaves/")
def list_leaves(
    employee_id: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(LeaveRequest)
    if employee_id:
        q = q.filter(LeaveRequest.employee_id == employee_id)
    if year:
        q = q.filter(sqlfunc.extract('year', LeaveRequest.date_from) == year)
    return [_leave_dict(l) for l in q.order_by(LeaveRequest.date_from.desc()).all()]


@router.post("/leaves/", status_code=201)
def create_leave(body: LeaveRequestCreate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    days = _business_days(body.date_from, body.date_to)
    leave = LeaveRequest(**body.model_dump(), days_count=days)
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return _leave_dict(leave)


@router.post("/leaves/{leave_id}/approve/")
def approve_leave(leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ("director", "projekt_leiter"):
        raise HTTPException(403, "Insufficient rights")
    leave = db.query(LeaveRequest).filter(LeaveRequest.id == leave_id).first()
    if not leave:
        raise HTTPException(404, "Not found")
    leave.status = "approved"
    leave.approved_by = current_user.id
    leave.approved_at = datetime.utcnow()
    # Auto-create time entries for the leave period
    cur = leave.date_from
    while cur <= leave.date_to:
        if cur.weekday() < 5:
            existing = db.query(TimeEntry).filter(
                TimeEntry.employee_id == leave.employee_id, TimeEntry.date == cur
            ).first()
            if not existing:
                db.add(TimeEntry(
                    employee_id=leave.employee_id,
                    date=cur,
                    hours_regular=0,
                    entry_type=leave.leave_type,
                    source="leave",
                    week_number=_iso_week(cur),
                    created_by=current_user.id,
                ))
        cur += timedelta(days=1)
    db.commit()
    return _leave_dict(leave)


@router.post("/leaves/{leave_id}/reject/")
def reject_leave(leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ("director", "projekt_leiter"):
        raise HTTPException(403, "Insufficient rights")
    leave = db.query(LeaveRequest).filter(LeaveRequest.id == leave_id).first()
    if not leave:
        raise HTTPException(404, "Not found")
    leave.status = "rejected"
    leave.approved_by = current_user.id
    leave.approved_at = datetime.utcnow()
    db.commit()
    return _leave_dict(leave)


# ── Payroll ───────────────────────────────────────────────────────────────────

@router.get("/payroll/")
def list_payroll(
    year: int = Query(...),
    month: int = Query(...),
    employee_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(PayrollRecord).filter(PayrollRecord.year == year, PayrollRecord.month == month)
    if employee_id:
        q = q.filter(PayrollRecord.employee_id == employee_id)
    return [_payroll_dict(p) for p in q.all()]


@router.post("/payroll/calculate/")
def calculate_payroll(
    body: PayrollRecalcRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Compute/recalculate payroll from time entries for a given month."""
    if current_user.role not in ("director", "projekt_leiter"):
        raise HTTPException(403, "Insufficient rights")

    emp_q = db.query(Employee).filter(Employee.is_active == True)
    if body.employee_ids:
        emp_q = emp_q.filter(Employee.id.in_(body.employee_ids))
    employees = emp_q.all()

    month_start = date(body.year, body.month, 1)
    if body.month == 12:
        month_end = date(body.year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(body.year, body.month + 1, 1) - timedelta(days=1)

    records = []
    for emp in employees:
        entries = db.query(TimeEntry).filter(
            TimeEntry.employee_id == emp.id,
            TimeEntry.date >= month_start,
            TimeEntry.date <= month_end,
        ).all()

        hours_regular  = sum(e.hours_regular for e in entries if e.entry_type == "work")
        hours_overtime = sum(e.hours_overtime for e in entries)
        hours_night    = sum(e.hours_night for e in entries)
        hours_holiday  = sum(e.hours_holiday for e in entries)
        hours_sick     = sum(1.0 for e in entries if e.entry_type == "sick") * 8
        hours_vacation = sum(1.0 for e in entries if e.entry_type == "vacation") * 8
        days_worked    = len([e for e in entries if e.entry_type == "work"])

        tariflohn    = emp.tariflohn or 0
        bauzuschlag  = emp.bauzuschlag or 0

        brutto_regular     = hours_regular * tariflohn
        brutto_overtime    = hours_overtime * tariflohn * 1.25
        brutto_night       = hours_night * tariflohn * 0.25
        brutto_bauzuschlag = (hours_regular + hours_overtime) * bauzuschlag
        brutto_total       = brutto_regular + brutto_overtime + brutto_night + brutto_bauzuschlag

        ag_sv_anteil = brutto_total * 0.214
        soka_bau     = brutto_total * 0.154
        total_employer_cost = brutto_total + ag_sv_anteil + soka_bau

        # Upsert
        existing = db.query(PayrollRecord).filter(
            PayrollRecord.employee_id == emp.id,
            PayrollRecord.year == body.year,
            PayrollRecord.month == body.month,
        ).first()

        if existing and existing.status == "locked":
            records.append(_payroll_dict(existing))
            continue

        data = dict(
            hours_regular=round(hours_regular, 2),
            hours_overtime=round(hours_overtime, 2),
            hours_night=round(hours_night, 2),
            hours_holiday=round(hours_holiday, 2),
            hours_sick=round(hours_sick, 2),
            hours_vacation=round(hours_vacation, 2),
            days_worked=days_worked,
            brutto_regular=round(brutto_regular, 2),
            brutto_overtime=round(brutto_overtime, 2),
            brutto_night=round(brutto_night, 2),
            brutto_bauzuschlag=round(brutto_bauzuschlag, 2),
            brutto_total=round(brutto_total, 2),
            ag_sv_anteil=round(ag_sv_anteil, 2),
            soka_bau=round(soka_bau, 2),
            total_employer_cost=round(total_employer_cost, 2),
        )
        if existing:
            for k, v in data.items():
                setattr(existing, k, v)
            records.append(_payroll_dict(existing))
        else:
            rec = PayrollRecord(employee_id=emp.id, year=body.year, month=body.month, **data)
            db.add(rec)
            db.flush()
            records.append(_payroll_dict(rec))

    db.commit()
    return records


@router.post("/payroll/{record_id}/lock/")
def lock_payroll(record_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != "director":
        raise HTTPException(403, "Only directors can lock payroll")
    rec = db.query(PayrollRecord).filter(PayrollRecord.id == record_id).first()
    if not rec:
        raise HTTPException(404, "Not found")
    rec.status = "locked"
    rec.locked_by = current_user.id
    rec.locked_at = datetime.utcnow()
    db.commit()
    return _payroll_dict(rec)


# ── DATEV Export ──────────────────────────────────────────────────────────────

@router.get("/payroll/datev-export/")
def datev_export(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export payroll as DATEV EXTF Buchungsstapel CSV."""
    if current_user.role not in ("director", "projekt_leiter"):
        raise HTTPException(403, "Insufficient rights")

    records = (
        db.query(PayrollRecord)
        .filter(PayrollRecord.year == year, PayrollRecord.month == month)
        .order_by(PayrollRecord.employee_id)
        .all()
    )

    if not records:
        raise HTTPException(404, f"Keine Abrechnungsdaten für {month:02d}/{year}")

    # Last day of month for Belegdatum
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    beleg_date = last_day.strftime("%d%m")  # DDMM format for DATEV

    now_str = datetime.utcnow().strftime("%Y%m%d%H%M%S000")

    buf = io.StringIO()
    # DATEV EXTF Vorlaufsatz (header line 1)
    buf.write(
        f'"EXTF";700;21;"Buchungsstapel";6;{now_str};;"";"";"";1000;'
        f'{year}0101;{last_day.strftime("%Y%m%d")};'
        f'"Lohnabrechnung {month:02d}/{year}";"AL";0;0;"EUR";"";"";"";"";;""'
        "\r\n"
    )
    # Column headers (line 2)
    buf.write(
        "Umsatz (ohne Soll/Haben-Kz);Soll/Haben-Kennzeichen;WKZ Umsatz;Kurs;"
        "Basis-Umsatz;WKZ Basis-Umsatz;Konto;Gegenkonto (ohne BU-Schlüssel);"
        "BU-Schlüssel;Belegdatum;Belegfeld 1;Belegfeld 2;Skonto;Buchungstext;"
        "Postensperre;Diverse Adressnummer;Geschäftspartnerbank;Sachverhalt;"
        "Zinssperre;Beleglink\r\n"
    )

    def row(betrag: float, konto: str, gegenkonto: str, text: str) -> str:
        return (
            f'"{betrag:.2f}".replace(".", ",");"S";"EUR";;;'
            f'"{konto}";"{gegenkonto}";;{beleg_date};"";"";;"{text}";;;;;;;'
            "\r\n"
        )

    MONTHS_DE = ["Januar","Februar","März","April","Mai","Juni",
                 "Juli","August","September","Oktober","November","Dezember"]

    for rec in records:
        emp = rec.employee
        name = f"{emp.nachname} {emp.vorname}" if emp else f"#{rec.employee_id}"
        pnr  = emp.personalnummer if emp and emp.personalnummer else str(rec.employee_id)
        monat = MONTHS_DE[month - 1]

        # Format amount with German decimal separator
        def fmt(val: float) -> str:
            return f"{val:.2f}".replace(".", ",")

        # 1. Bruttolohn → Konto 4120 (Löhne Arbeiter) / Gegenkonto 3720 (Verbindlichk. Lohn)
        if rec.brutto_total > 0:
            buf.write(
                f'{fmt(rec.brutto_total)};S;EUR;;;4120;3720;;{beleg_date};{pnr};;'
                f';"Lohn {name} {monat} {year}";;;;;;;\r\n'
            )
        # 2. AG-SV-Anteil → Konto 4130 / Gegenkonto 3720
        if rec.ag_sv_anteil > 0:
            buf.write(
                f'{fmt(rec.ag_sv_anteil)};S;EUR;;;4130;3720;;{beleg_date};{pnr};;'
                f';"AG-SV {name} {monat} {year}";;;;;;;\r\n'
            )
        # 3. SOKA-BAU → Konto 4160 / Gegenkonto 3720
        if rec.soka_bau > 0:
            buf.write(
                f'{fmt(rec.soka_bau)};S;EUR;;;4160;3720;;{beleg_date};{pnr};;'
                f';"SOKA-BAU {name} {monat} {year}";;;;;;;\r\n'
            )

    csv_bytes = buf.getvalue().encode("latin-1", errors="replace")
    filename = f"DATEV_Lohn_{year}_{month:02d}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=latin-1",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Team Assignments ──────────────────────────────────────────────────────────

from app.models.timesheet import TeamAssignment


class TeamAssignmentSet(BaseModel):
    employee_ids: List[int]


@router.get("/team-assignments/")
def list_team_assignments(
    team_lead_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Return all assignments. Optional filter by team_lead_id."""
    q = db.query(TeamAssignment)
    if team_lead_id:
        q = q.filter(TeamAssignment.team_lead_id == team_lead_id)
    result = []
    for a in q.all():
        emp = a.employee
        result.append({
            "id": a.id,
            "team_lead_id": a.team_lead_id,
            "team_lead_name": f"{a.team_lead.full_name}" if a.team_lead else None,
            "employee_id": a.employee_id,
            "employee_name": f"{emp.vorname} {emp.nachname}" if emp else None,
        })
    return result


@router.put("/team-assignments/{team_lead_id}/")
def set_team_assignments(
    team_lead_id: int,
    body: TeamAssignmentSet,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Replace all assignments for a team lead. Director only."""
    if current_user.role not in ("director", "projekt_leiter"):
        raise HTTPException(403, "Insufficient rights")
    # Delete existing
    db.query(TeamAssignment).filter(TeamAssignment.team_lead_id == team_lead_id).delete()
    # Insert new
    for emp_id in body.employee_ids:
        db.add(TeamAssignment(team_lead_id=team_lead_id, employee_id=emp_id))
    db.commit()
    return {"team_lead_id": team_lead_id, "assigned": len(body.employee_ids)}


@router.get("/my-team/")
def my_team(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return employees assigned to the current user (mobile team lead)."""
    assignments = db.query(TeamAssignment).filter(
        TeamAssignment.team_lead_id == current_user.id
    ).all()
    return [
        {
            "employee_id": a.employee_id,
            "employee_name": f"{a.employee.vorname} {a.employee.nachname}" if a.employee else None,
        }
        for a in assignments
    ]


# ── Mobile Pontaj bulk submit ─────────────────────────────────────────────────

class PontajWorkerIn(BaseModel):
    employee_id: int
    present: bool
    ora_start: Optional[str] = None
    ora_stop: Optional[str] = None
    notes: Optional[str] = None


class PontajSubmit(BaseModel):
    local_uuid: str
    date: date
    site_id: Optional[int] = None
    workers: List[PontajWorkerIn]


@router.post("/pontaj/", status_code=201)
def submit_pontaj(
    body: PontajSubmit,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Bulk upsert time entries submitted from mobile pontaj screen."""
    saved = 0
    for w in body.workers:
        entry_type = "work" if w.present else "absent"
        # compute hours from start/stop if present
        hours = 8.0
        if w.present and w.ora_start and w.ora_stop:
            try:
                sh, sm = map(int, w.ora_start.split(":"))
                eh, em = map(int, w.ora_stop.split(":"))
                hours = round((eh * 60 + em - sh * 60 - sm) / 60, 2)
                if hours < 0:
                    hours = 0.0
            except Exception:
                hours = 8.0

        existing = db.query(TimeEntry).filter(
            TimeEntry.employee_id == w.employee_id,
            TimeEntry.date == body.date,
        ).first()

        if existing:
            existing.entry_type = entry_type
            existing.hours_regular = hours if w.present else 0.0
            existing.ora_start = w.ora_start
            existing.ora_stop = w.ora_stop
            existing.notes = w.notes
            existing.site_id = body.site_id
            existing.team_lead_id = current_user.id
            existing.source = "pontaj_mobile"
            existing.week_number = _iso_week(body.date)
        else:
            db.add(TimeEntry(
                employee_id=w.employee_id,
                site_id=body.site_id,
                date=body.date,
                hours_regular=hours if w.present else 0.0,
                entry_type=entry_type,
                ora_start=w.ora_start,
                ora_stop=w.ora_stop,
                notes=w.notes,
                team_lead_id=current_user.id,
                source="pontaj_mobile",
                week_number=_iso_week(body.date),
                created_by=current_user.id,
            ))
        saved += 1

    db.commit()
    return {"saved": saved, "date": str(body.date)}


@router.get("/pontaj/")
def list_pontaj(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    team_lead_id: Optional[int] = Query(None),
    site_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """List pontaj entries (time entries from mobile source) for HestiOS view."""
    q = db.query(TimeEntry).filter(TimeEntry.source == "pontaj_mobile")
    if date_from:
        q = q.filter(TimeEntry.date >= date_from)
    if date_to:
        q = q.filter(TimeEntry.date <= date_to)
    if team_lead_id:
        q = q.filter(TimeEntry.team_lead_id == team_lead_id)
    if site_id:
        q = q.filter(TimeEntry.site_id == site_id)
    entries = q.order_by(TimeEntry.date.desc(), TimeEntry.employee_id).all()
    result = []
    for e in entries:
        emp = e.employee
        result.append({
            "id": e.id,
            "employee_id": e.employee_id,
            "employee_name": f"{emp.vorname} {emp.nachname}" if emp else None,
            "site_id": e.site_id,
            "date": str(e.date),
            "entry_type": e.entry_type,
            "hours_regular": e.hours_regular,
            "ora_start": e.ora_start,
            "ora_stop": e.ora_stop,
            "notes": e.notes,
            "team_lead_id": e.team_lead_id,
        })
    return result
