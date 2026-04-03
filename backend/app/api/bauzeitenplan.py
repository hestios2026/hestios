from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, timedelta
import io

from app.core.database import get_db
from app.models.bauzeitenplan import BauzeitenplanProject, BauzeitenplanRow, BauzeitenplanWeekly
from app.models.site import Site
from app.models.user import User
from app.api.auth import get_current_user

router = APIRouter(prefix="/bauzeitenplan", tags=["bauzeitenplan"])

ALLOWED_ROLES = ("director", "projekt_leiter", "aufmass")


# ── Serializers ──────────────────────────────────────────────────────────────

def _weekly_dict(w: BauzeitenplanWeekly):
    return {
        "id": w.id,
        "row_id": w.row_id,
        "week_date": str(w.week_date),
        "meters": w.meters,
        "note": w.note,
    }


def _row_dict(r: BauzeitenplanRow, include_weekly=True):
    d = {
        "id": r.id,
        "project_id": r.project_id,
        "vorhaben_nr": r.vorhaben_nr,
        "hk_nvt": r.hk_nvt,
        "gewerk": r.gewerk,
        "hh": r.hh,
        "hc": r.hc,
        "tb_soll_m": r.tb_soll_m,
        "date_start": str(r.date_start) if r.date_start else None,
        "date_end": str(r.date_end) if r.date_end else None,
        "tb_ist_m": r.tb_ist_m,
        "ha_gebaut": r.ha_gebaut,
        "verzug_kw": r.verzug_kw,
        "bemerkung": r.bemerkung,
        "sort_order": r.sort_order,
        "is_group_header": r.is_group_header,
        "color": r.color,
        "progress_pct": r.progress_pct,
    }
    if include_weekly:
        d["weekly"] = [_weekly_dict(w) for w in r.weekly]
    return d


def _project_dict(p: BauzeitenplanProject, include_rows=True):
    d = {
        "id": p.id,
        "site_id": p.site_id,
        "name": p.name,
        "firma": p.firma,
        "baubeginn": str(p.baubeginn) if p.baubeginn else None,
        "bauende": str(p.bauende) if p.bauende else None,
        "site_name": p.site.name if p.site else None,
    }
    if include_rows:
        d["rows"] = [_row_dict(r) for r in p.rows]
    return d


# ── Schemas ──────────────────────────────────────────────────────────────────

class ProjectCreate(BaseModel):
    site_id: int
    name: str
    firma: Optional[str] = None
    baubeginn: Optional[date] = None
    bauende: Optional[date] = None


class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    firma: Optional[str] = None
    baubeginn: Optional[date] = None
    bauende: Optional[date] = None


class RowCreate(BaseModel):
    vorhaben_nr: Optional[str] = None
    hk_nvt: Optional[str] = None
    gewerk: Optional[str] = None
    hh: bool = False
    hc: bool = False
    tb_soll_m: Optional[float] = None
    date_start: Optional[date] = None
    date_end: Optional[date] = None
    tb_ist_m: float = 0
    ha_gebaut: int = 0
    verzug_kw: int = 0
    bemerkung: Optional[str] = None
    sort_order: int = 0
    is_group_header: bool = False
    color: Optional[str] = None


class RowUpdate(BaseModel):
    vorhaben_nr: Optional[str] = None
    hk_nvt: Optional[str] = None
    gewerk: Optional[str] = None
    hh: Optional[bool] = None
    hc: Optional[bool] = None
    tb_soll_m: Optional[float] = None
    date_start: Optional[date] = None
    date_end: Optional[date] = None
    tb_ist_m: Optional[float] = None
    ha_gebaut: Optional[int] = None
    verzug_kw: Optional[int] = None
    bemerkung: Optional[str] = None
    sort_order: Optional[int] = None
    is_group_header: Optional[bool] = None
    color: Optional[str] = None


class WeeklyUpsert(BaseModel):
    week_date: date
    meters: float
    note: Optional[str] = None


class RowsReorder(BaseModel):
    row_ids: List[int]  # ordered list


# ── Projects ─────────────────────────────────────────────────────────────────

@router.get("/projects/")
def list_projects(
    site_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(BauzeitenplanProject)
    if site_id:
        q = q.filter(BauzeitenplanProject.site_id == site_id)
    return [_project_dict(p, include_rows=False) for p in q.order_by(BauzeitenplanProject.id.desc()).all()]


@router.get("/projects/{project_id}/")
def get_project(
    project_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    p = db.query(BauzeitenplanProject).filter(BauzeitenplanProject.id == project_id).first()
    if not p:
        raise HTTPException(404, "Not found")
    return _project_dict(p)


@router.post("/projects/", status_code=201)
def create_project(
    body: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    p = BauzeitenplanProject(**body.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return _project_dict(p)


@router.put("/projects/{project_id}/")
def update_project(
    project_id: int,
    body: ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    p = db.query(BauzeitenplanProject).filter(BauzeitenplanProject.id == project_id).first()
    if not p:
        raise HTTPException(404, "Not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    return _project_dict(p)


@router.delete("/projects/{project_id}/", status_code=204)
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("director", "projekt_leiter"):
        raise HTTPException(403, "Insufficient rights")
    p = db.query(BauzeitenplanProject).filter(BauzeitenplanProject.id == project_id).first()
    if not p:
        raise HTTPException(404, "Not found")
    db.delete(p)
    db.commit()


# ── Rows ─────────────────────────────────────────────────────────────────────

@router.post("/projects/{project_id}/rows/", status_code=201)
def add_row(
    project_id: int,
    body: RowCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    p = db.query(BauzeitenplanProject).filter(BauzeitenplanProject.id == project_id).first()
    if not p:
        raise HTTPException(404, "Project not found")
    row = BauzeitenplanRow(project_id=project_id, **body.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return _row_dict(row)


@router.put("/rows/{row_id}/")
def update_row(
    row_id: int,
    body: RowUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    row = db.query(BauzeitenplanRow).filter(BauzeitenplanRow.id == row_id).first()
    if not row:
        raise HTTPException(404, "Not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(row, k, v)
    # recalculate tb_ist_m from weekly if updated via weekly
    db.commit()
    db.refresh(row)
    return _row_dict(row)


@router.delete("/rows/{row_id}/", status_code=204)
def delete_row(
    row_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    row = db.query(BauzeitenplanRow).filter(BauzeitenplanRow.id == row_id).first()
    if not row:
        raise HTTPException(404, "Not found")
    db.delete(row)
    db.commit()


@router.post("/projects/{project_id}/reorder/")
def reorder_rows(
    project_id: int,
    body: RowsReorder,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    for order, row_id in enumerate(body.row_ids):
        db.query(BauzeitenplanRow).filter(
            BauzeitenplanRow.id == row_id,
            BauzeitenplanRow.project_id == project_id,
        ).update({"sort_order": order})
    db.commit()
    return {"ok": True}


# ── Weekly values ────────────────────────────────────────────────────────────

@router.put("/rows/{row_id}/weekly/")
def upsert_weekly(
    row_id: int,
    body: WeeklyUpsert,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ALLOWED_ROLES:
        raise HTTPException(403, "Insufficient rights")
    row = db.query(BauzeitenplanRow).filter(BauzeitenplanRow.id == row_id).first()
    if not row:
        raise HTTPException(404, "Row not found")

    existing = db.query(BauzeitenplanWeekly).filter(
        BauzeitenplanWeekly.row_id == row_id,
        BauzeitenplanWeekly.week_date == body.week_date,
    ).first()

    if existing:
        existing.meters = body.meters
        existing.note = body.note
    else:
        db.add(BauzeitenplanWeekly(row_id=row_id, **body.model_dump()))

    # Update tb_ist_m = sum of all weekly meters
    db.flush()
    total = db.query(BauzeitenplanWeekly).filter(BauzeitenplanWeekly.row_id == row_id).all()
    row.tb_ist_m = sum(w.meters for w in total)
    db.commit()
    db.refresh(row)
    return _row_dict(row)


# ── Export Excel ─────────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/export/excel/")
def export_excel(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    p = db.query(BauzeitenplanProject).filter(BauzeitenplanProject.id == project_id).first()
    if not p:
        raise HTTPException(404, "Not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bauzeitenplan"

    # Color map
    COLORS = {
        "Tiefbau": "F97316",
        "Montage": "3B82F6",
        "Spülbohrung": "8B5CF6",
        "Einblasen": "10B981",
        "Rohreinzug": "F59E0B",
        "header": "1E293B",
    }

    # Collect all KW dates across all rows
    all_weeks = set()
    for row in p.rows:
        for w in row.weekly:
            all_weeks.add(w.week_date)
        if row.date_start and row.date_end:
            cur = row.date_start
            while cur <= row.date_end:
                all_weeks.add(cur - timedelta(days=cur.weekday()))
                cur += timedelta(weeks=1)
    sorted_weeks = sorted(all_weeks)

    # Header row 1 — project info
    ws.cell(1, 1, "Bauzeitenplan:").font = Font(bold=True)
    ws.cell(1, 2, p.name).font = Font(bold=True)
    ws.cell(2, 1, "Firma:").font = Font(bold=True)
    ws.cell(2, 2, p.firma or "")
    ws.cell(3, 1, "Baubeginn:").font = Font(bold=True)
    ws.cell(3, 2, str(p.baubeginn) if p.baubeginn else "")

    # Column headers
    FIXED_COLS = ["Vorhaben", "HK/NVT", "Gewerk", "HH", "HC",
                  "Tb Soll (m)", "Baubeginn", "Bauende",
                  "Tb Ist (m)", "Fortschritt (%)", "HA gebaut", "Verzug KW", "Bemerkung"]
    HDR_ROW = 5
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(FIXED_COLS, 1):
        c = ws.cell(HDR_ROW, ci, h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    kw_start_col = len(FIXED_COLS) + 1
    for ki, wdate in enumerate(sorted_weeks):
        ci = kw_start_col + ki
        iso = wdate.isocalendar()
        c = ws.cell(HDR_ROW, ci, f"KW{iso[1]}\n{wdate.strftime('%d.%m')}")
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = 6

    # Data rows
    for ri, row in enumerate(p.rows):
        dr = HDR_ROW + 1 + ri
        gewerk = (row.gewerk or "").strip()
        color_hex = row.color or COLORS.get(gewerk, "E2E8F0")
        row_fill = PatternFill("solid", fgColor=color_hex) if not row.is_group_header else PatternFill("solid", fgColor="334155")

        def wc(col, val):
            c = ws.cell(dr, col, val)
            c.fill = row_fill
            c.font = Font(color="FFFFFF" if row.is_group_header else "1E293B", size=9)
            c.alignment = Alignment(vertical="center")
            c.border = border
            return c

        wc(1, row.vorhaben_nr or "")
        wc(2, row.hk_nvt or "")
        wc(3, gewerk)
        wc(4, "x" if row.hh else "")
        wc(5, "x" if row.hc else "")
        wc(6, row.tb_soll_m or "")
        wc(7, str(row.date_start) if row.date_start else "")
        wc(8, str(row.date_end) if row.date_end else "")
        wc(9, row.tb_ist_m or 0)
        wc(10, f"{row.progress_pct}%" if row.tb_soll_m else "")
        wc(11, row.ha_gebaut or "")
        wc(12, row.verzug_kw or "")
        wc(13, row.bemerkung or "")

        # Weekly columns
        weekly_map = {w.week_date: w for w in row.weekly}
        for ki, wdate in enumerate(sorted_weeks):
            ci = kw_start_col + ki
            w_entry = weekly_map.get(wdate)
            val = w_entry.meters if w_entry and w_entry.meters else ""
            note = w_entry.note if w_entry else ""
            cell = ws.cell(dr, ci, val if val else "")
            if val and not row.is_group_header:
                cell.fill = PatternFill("solid", fgColor=color_hex)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if note:
                cell.comment = None  # openpyxl comment support limited

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 4
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 20
    ws.row_dimensions[HDR_ROW].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"BZP_{p.name.replace(' ','_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Export PDF ───────────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/export/pdf/")
def export_pdf(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    p = db.query(BauzeitenplanProject).filter(BauzeitenplanProject.id == project_id).first()
    if not p:
        raise HTTPException(404, "Not found")

    # Collect weeks
    all_weeks = set()
    for row in p.rows:
        for w in row.weekly:
            all_weeks.add(w.week_date)
        if row.date_start and row.date_end:
            cur = row.date_start
            while cur <= row.date_end:
                all_weeks.add(cur - timedelta(days=cur.weekday()))
                cur += timedelta(weeks=1)
    sorted_weeks = sorted(all_weeks)

    GEWERK_COLORS = {
        "Tiefbau":     colors.HexColor("#F97316"),
        "Montage":     colors.HexColor("#3B82F6"),
        "Spülbohrung": colors.HexColor("#8B5CF6"),
        "Einblasen":   colors.HexColor("#10B981"),
        "Rohreinzug":  colors.HexColor("#F59E0B"),
    }
    HDR_COLOR = colors.HexColor("#1E293B")
    WHITE = colors.white

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3), leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"<b>Bauzeitenplan: {p.name}</b>", styles["Title"]))
    if p.firma:
        story.append(Paragraph(f"Firma: {p.firma}", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    # Table header
    fixed_headers = ["HK/NVT", "Gewerk", "Tb Soll", "Start", "Ende", "Tb Ist", "%", "Bemerkung"]
    kw_headers = [f"KW{w.isocalendar()[1]}" for w in sorted_weeks]
    header_row = fixed_headers + kw_headers

    table_data = [header_row]
    table_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), HDR_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]

    for ri, row in enumerate(p.rows, 1):
        gewerk = (row.gewerk or "").strip()
        weekly_map = {w.week_date: w for w in row.weekly}

        # Has this row any active weeks?
        kw_cells = []
        for wdate in sorted_weeks:
            w_entry = weekly_map.get(wdate)
            val = str(int(w_entry.meters)) if w_entry and w_entry.meters else ""
            kw_cells.append(val)

        data_row = [
            row.hk_nvt or "",
            gewerk,
            f"{row.tb_soll_m:.0f}m" if row.tb_soll_m else "",
            str(row.date_start) if row.date_start else "",
            str(row.date_end) if row.date_end else "",
            f"{row.tb_ist_m:.0f}m" if row.tb_ist_m else "",
            f"{row.progress_pct}%" if row.tb_soll_m else "",
            row.bemerkung or "",
        ] + kw_cells

        table_data.append(data_row)

        if row.is_group_header:
            table_styles.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#334155")))
            table_styles.append(("TEXTCOLOR", (0, ri), (-1, ri), WHITE))
            table_styles.append(("FONTNAME", (0, ri), (-1, ri), "Helvetica-Bold"))
        else:
            row_color = GEWERK_COLORS.get(gewerk)
            if row_color:
                # Color only the KW columns that have values
                for ki, wdate in enumerate(sorted_weeks):
                    w_entry = weekly_map.get(wdate)
                    if w_entry and w_entry.meters:
                        ci = len(fixed_headers) + ki
                        table_styles.append(("BACKGROUND", (ci, ri), (ci, ri), row_color))
                        table_styles.append(("TEXTCOLOR", (ci, ri), (ci, ri), WHITE))

    # Column widths
    fixed_widths = [2.2*cm, 2*cm, 1.5*cm, 2*cm, 2*cm, 1.5*cm, 1.2*cm, 3*cm]
    kw_widths = [0.85*cm] * len(sorted_weeks)
    col_widths = fixed_widths + kw_widths

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(table_styles))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    fname = f"BZP_{p.name.replace(' ','_')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
